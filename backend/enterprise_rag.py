"""Versioned, auditable standard-knowledge RAG for M-AI Master.

The module is intentionally independent from Flask.  ``EnterpriseRAG`` can be
constructed with the application's SQLite path and used by routes without
changing the current response contract.  XLSX import prefers openpyxl when it
is installed and otherwise uses a small, read-only OOXML parser from the
standard library.
"""

from __future__ import annotations

import hashlib
import json
import math
import posixpath
import re
import sqlite3
import struct
import uuid
import zipfile
from contextlib import contextmanager
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, Mapping, Sequence
from xml.etree import ElementTree as ET


STATUS_DRAFT = "DRAFT"
STATUS_VALIDATED = "VALIDATED"
STATUS_PUBLISHED = "PUBLISHED"
STATUS_RETIRED = "RETIRED"
VERSION_STATUSES = {STATUS_DRAFT, STATUS_VALIDATED, STATUS_PUBLISHED, STATUS_RETIRED}

CLASSIFICATION_SHEET = "备品备件分类树"
MAPPING_SHEET = "物料归类映射表"
REFERENCE_SHEET = "标准选用规范"

CLASSIFICATION_HEADERS = (
    "大类编码", "大类名称", "中类编码", "中类名称", "小类编码", "小类名称",
    "品名编码", "品名名称", "物料业务类型", "SY/T5497完整8位编码", "适配物料示例",
)
MAPPING_HEADERS = ("物料名称关键词", "归属大类编码", "归属大类名称", "推荐8位编码", "适配层级说明")
REFERENCE_HEADERS = ("适用场景", "标准编号", "标准名称", "作用")

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_RE = re.compile(r"^([A-Z]+)(\d+)$")
_CODE_RE = re.compile(r"^\d+$")
_SPLIT_RE = re.compile(r"[\s、,，/；;|]+")
_ALNUM_RE = re.compile(r"[a-z0-9]+(?:[-./][a-z0-9]+)*", re.I)


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_text(value: str) -> str:
    return _sha256_bytes(value.encode("utf-8"))


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def _json_list(value: Sequence[str] | None, default: Sequence[str] = ("*",)) -> str:
    cleaned = list(dict.fromkeys(_clean(item).upper() for item in (value or default) if _clean(item)))
    return _canonical_json(cleaned or list(default))


def _parse_json(value: str | None, default: Any) -> Any:
    try:
        return json.loads(value or "")
    except (TypeError, json.JSONDecodeError):
        return default


def _column_number(reference: str) -> int:
    match = _CELL_RE.match(reference)
    if not match:
        raise ValueError(f"invalid XLSX cell reference: {reference}")
    number = 0
    for char in match.group(1):
        number = number * 26 + ord(char) - 64
    return number


def _column_name(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


@dataclass(frozen=True)
class RAGConfig:
    """Retrieval settings persisted as named profiles."""

    dimension: int = 384
    rule_weight: float = 0.45
    lexical_weight: float = 0.30
    vector_weight: float = 0.25
    auto_accept_threshold: float = 0.68
    minimum_margin: float = 0.08
    default_top_k: int = 3
    audit_queries: bool = True

    def validate(self) -> "RAGConfig":
        if not 32 <= int(self.dimension) <= 4096:
            raise ValueError("dimension must be between 32 and 4096")
        weights = (self.rule_weight, self.lexical_weight, self.vector_weight)
        if any(weight < 0 for weight in weights) or sum(weights) <= 0:
            raise ValueError("retrieval weights must be non-negative and not all zero")
        if not 0 <= self.auto_accept_threshold <= 1:
            raise ValueError("auto_accept_threshold must be between 0 and 1")
        if not 0 <= self.minimum_margin <= 1:
            raise ValueError("minimum_margin must be between 0 and 1")
        if not 1 <= int(self.default_top_k) <= 50:
            raise ValueError("default_top_k must be between 1 and 50")
        return self

    @classmethod
    def from_mapping(cls, value: Mapping[str, Any] | None) -> "RAGConfig":
        if not value:
            return cls().validate()
        allowed = set(cls.__dataclass_fields__)
        return cls(**{key: item for key, item in value.items() if key in allowed}).validate()

    def normalized_weights(self) -> tuple[float, float, float]:
        total = self.rule_weight + self.lexical_weight + self.vector_weight
        return self.rule_weight / total, self.lexical_weight / total, self.vector_weight / total


class EnterpriseRAG:
    """SQLite-backed, versioned knowledge base with hybrid retrieval.

    Public methods are framework-neutral.  The compatible methods most useful
    to Flask routes are ``stats()``, ``search()`` and ``classify()``.
    Administrative lifecycle methods are ``import_xlsx()``,
    ``validate_version()``, ``publish_version()`` and ``rollback_version()``.
    """

    def __init__(
        self,
        db_path: str | Path,
        *,
        catalog_id: str = "syt5497",
        catalog_name: str = "SY/T 5497 石油工业物资分类知识库",
        standard_no: str = "SY/T 5497-2018",
        config: RAGConfig | Mapping[str, Any] | None = None,
        embedding_fn: Callable[[str, int], Sequence[float]] | None = None,
        max_source_bytes: int = 25 * 1024 * 1024,
        max_uncompressed_bytes: int = 100 * 1024 * 1024,
    ) -> None:
        self.db_path = str(db_path)
        self.catalog_id = _clean(catalog_id) or "syt5497"
        self.catalog_name = _clean(catalog_name)
        self.standard_no = _clean(standard_no)
        self.config = config if isinstance(config, RAGConfig) else RAGConfig.from_mapping(config)
        self.config.validate()
        self.embedding_fn = embedding_fn
        self.max_source_bytes = int(max_source_bytes)
        self.max_uncompressed_bytes = int(max_uncompressed_bytes)
        self._initialize_schema()

    @contextmanager
    def _connect(self) -> Iterator[sqlite3.Connection]:
        connection = sqlite3.connect(self.db_path, timeout=30)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        connection.execute("PRAGMA busy_timeout = 30000")
        try:
            yield connection
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

    @contextmanager
    def _write(self) -> Iterator[sqlite3.Connection]:
        with self._connect() as connection:
            connection.execute("BEGIN IMMEDIATE")
            yield connection

    def _initialize_schema(self) -> None:
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS enterprise_rag_catalogs (
                    catalog_id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    standard_no TEXT NOT NULL,
                    active_version_id TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_sources (
                    source_id TEXT PRIMARY KEY,
                    source_sha256 TEXT NOT NULL UNIQUE,
                    filename TEXT NOT NULL,
                    source_path TEXT,
                    byte_size INTEGER NOT NULL,
                    imported_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_versions (
                    version_id TEXT PRIMARY KEY,
                    catalog_id TEXT NOT NULL,
                    version_label TEXT NOT NULL,
                    status TEXT NOT NULL CHECK(status IN ('DRAFT','VALIDATED','PUBLISHED','RETIRED')),
                    source_id TEXT NOT NULL,
                    source_sha256 TEXT NOT NULL,
                    entry_count INTEGER NOT NULL DEFAULT 0,
                    alias_count INTEGER NOT NULL DEFAULT 0,
                    reference_count INTEGER NOT NULL DEFAULT 0,
                    validation_json TEXT,
                    allowed_plants_json TEXT NOT NULL,
                    allowed_classifications_json TEXT NOT NULL,
                    security_classification TEXT NOT NULL DEFAULT 'INTERNAL',
                    notes TEXT,
                    created_by TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    validated_at TEXT,
                    published_at TEXT,
                    retired_at TEXT,
                    FOREIGN KEY (catalog_id) REFERENCES enterprise_rag_catalogs(catalog_id),
                    FOREIGN KEY (source_id) REFERENCES enterprise_rag_sources(source_id),
                    UNIQUE(catalog_id, version_label),
                    UNIQUE(catalog_id, source_sha256)
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_entries (
                    entry_id TEXT PRIMARY KEY,
                    version_id TEXT NOT NULL,
                    major_code TEXT NOT NULL,
                    major_name TEXT NOT NULL,
                    middle_code TEXT NOT NULL,
                    middle_name TEXT NOT NULL,
                    minor_code TEXT NOT NULL,
                    minor_name TEXT NOT NULL,
                    item_code TEXT NOT NULL,
                    item_name TEXT NOT NULL,
                    business_type TEXT NOT NULL,
                    full_code TEXT NOT NULL,
                    example TEXT NOT NULL,
                    content TEXT NOT NULL,
                    content_hash TEXT NOT NULL,
                    source_sheet TEXT NOT NULL,
                    source_row INTEGER NOT NULL,
                    source_range TEXT NOT NULL,
                    row_sha256 TEXT NOT NULL,
                    security_classification TEXT NOT NULL,
                    plant_acl_json TEXT NOT NULL,
                    FOREIGN KEY (version_id) REFERENCES enterprise_rag_versions(version_id) ON DELETE CASCADE,
                    UNIQUE(version_id, full_code)
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_aliases (
                    alias_id TEXT PRIMARY KEY,
                    version_id TEXT NOT NULL,
                    keywords TEXT NOT NULL,
                    major_code TEXT NOT NULL,
                    major_name TEXT NOT NULL,
                    target_code TEXT NOT NULL,
                    level_note TEXT NOT NULL,
                    target_scope TEXT NOT NULL CHECK(target_scope IN ('IN_TREE','EXTERNAL_BOUNDARY')),
                    source_sheet TEXT NOT NULL,
                    source_row INTEGER NOT NULL,
                    source_range TEXT NOT NULL,
                    row_sha256 TEXT NOT NULL,
                    FOREIGN KEY (version_id) REFERENCES enterprise_rag_versions(version_id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_references (
                    reference_id TEXT PRIMARY KEY,
                    version_id TEXT NOT NULL,
                    scenario TEXT NOT NULL,
                    standard_no TEXT NOT NULL,
                    standard_name TEXT NOT NULL,
                    purpose TEXT NOT NULL,
                    source_sheet TEXT NOT NULL,
                    source_row INTEGER NOT NULL,
                    source_range TEXT NOT NULL,
                    row_sha256 TEXT NOT NULL,
                    FOREIGN KEY (version_id) REFERENCES enterprise_rag_versions(version_id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_chunks (
                    chunk_id TEXT PRIMARY KEY,
                    version_id TEXT NOT NULL,
                    entity_type TEXT NOT NULL,
                    entity_id TEXT NOT NULL,
                    content TEXT NOT NULL,
                    content_hash TEXT NOT NULL,
                    source_sheet TEXT NOT NULL,
                    source_row INTEGER NOT NULL,
                    source_range TEXT NOT NULL,
                    row_sha256 TEXT NOT NULL,
                    metadata_json TEXT NOT NULL,
                    security_classification TEXT NOT NULL,
                    plant_acl_json TEXT NOT NULL,
                    FOREIGN KEY (version_id) REFERENCES enterprise_rag_versions(version_id) ON DELETE CASCADE,
                    UNIQUE(version_id, entity_type, entity_id)
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_embeddings (
                    chunk_id TEXT NOT NULL,
                    provider TEXT NOT NULL,
                    model TEXT NOT NULL,
                    dimension INTEGER NOT NULL,
                    vector BLOB NOT NULL,
                    content_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    PRIMARY KEY (chunk_id, provider, model),
                    FOREIGN KEY (chunk_id) REFERENCES enterprise_rag_chunks(chunk_id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_profiles (
                    catalog_id TEXT NOT NULL,
                    profile_name TEXT NOT NULL,
                    config_json TEXT NOT NULL,
                    is_active INTEGER NOT NULL DEFAULT 0,
                    updated_by TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    PRIMARY KEY (catalog_id, profile_name),
                    FOREIGN KEY (catalog_id) REFERENCES enterprise_rag_catalogs(catalog_id)
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_query_logs (
                    query_id TEXT PRIMARY KEY,
                    catalog_id TEXT NOT NULL,
                    version_id TEXT NOT NULL,
                    query_text TEXT NOT NULL,
                    query_sha256 TEXT NOT NULL,
                    plant_code TEXT NOT NULL,
                    clearance TEXT NOT NULL,
                    profile_name TEXT NOT NULL,
                    results_json TEXT NOT NULL,
                    trace_id TEXT,
                    actor TEXT,
                    created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS enterprise_rag_audit_events (
                    event_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    catalog_id TEXT NOT NULL,
                    version_id TEXT,
                    event_type TEXT NOT NULL,
                    actor TEXT NOT NULL,
                    payload_json TEXT NOT NULL,
                    previous_hash TEXT NOT NULL,
                    event_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_enterprise_rag_versions_catalog
                    ON enterprise_rag_versions(catalog_id, status, created_at);
                CREATE UNIQUE INDEX IF NOT EXISTS idx_enterprise_rag_one_published
                    ON enterprise_rag_versions(catalog_id) WHERE status = 'PUBLISHED';
                CREATE INDEX IF NOT EXISTS idx_enterprise_rag_entries_version
                    ON enterprise_rag_entries(version_id, full_code);
                CREATE INDEX IF NOT EXISTS idx_enterprise_rag_aliases_version
                    ON enterprise_rag_aliases(version_id, target_code);
                CREATE INDEX IF NOT EXISTS idx_enterprise_rag_chunks_version
                    ON enterprise_rag_chunks(version_id, entity_type);
                CREATE INDEX IF NOT EXISTS idx_enterprise_rag_queries_version
                    ON enterprise_rag_query_logs(version_id, created_at);
                CREATE INDEX IF NOT EXISTS idx_enterprise_rag_audit_catalog
                    ON enterprise_rag_audit_events(catalog_id, event_id);
                """
            )
            now = _utc_now()
            connection.execute(
                """INSERT INTO enterprise_rag_catalogs
                   (catalog_id, name, standard_no, active_version_id, created_at, updated_at)
                   VALUES (?, ?, ?, NULL, ?, ?)
                   ON CONFLICT(catalog_id) DO UPDATE SET
                    name=excluded.name, standard_no=excluded.standard_no, updated_at=excluded.updated_at""",
                (self.catalog_id, self.catalog_name, self.standard_no, now, now),
            )
            if not connection.execute(
                "SELECT 1 FROM enterprise_rag_profiles WHERE catalog_id = ? LIMIT 1", (self.catalog_id,)
            ).fetchone():
                connection.execute(
                    """INSERT INTO enterprise_rag_profiles
                       (catalog_id, profile_name, config_json, is_active, updated_by, updated_at)
                       VALUES (?, 'default', ?, 1, 'system', ?)""",
                    (self.catalog_id, _canonical_json(asdict(self.config)), now),
                )

    def put_profile(
        self, profile_name: str, config: RAGConfig | Mapping[str, Any], *, actor: str = "admin", activate: bool = False
    ) -> dict[str, Any]:
        resolved = config if isinstance(config, RAGConfig) else RAGConfig.from_mapping(config)
        resolved.validate()
        name = _clean(profile_name)
        if not name:
            raise ValueError("profile_name is required")
        with self._write() as connection:
            if activate:
                connection.execute("UPDATE enterprise_rag_profiles SET is_active = 0 WHERE catalog_id = ?", (self.catalog_id,))
            connection.execute(
                """INSERT INTO enterprise_rag_profiles
                   (catalog_id, profile_name, config_json, is_active, updated_by, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?)
                   ON CONFLICT(catalog_id, profile_name) DO UPDATE SET
                    config_json=excluded.config_json, is_active=excluded.is_active,
                    updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
                (self.catalog_id, name, _canonical_json(asdict(resolved)), int(activate), _clean(actor) or "admin", _utc_now()),
            )
            self._append_audit(connection, "RAG_PROFILE_UPDATED", None, actor, {
                "profile_name": name, "active": bool(activate), "config": asdict(resolved),
            })
        return {"profile_name": name, "active": bool(activate), "config": asdict(resolved)}

    def _profile(self, profile_name: str | None) -> tuple[str, RAGConfig]:
        with self._connect() as connection:
            if profile_name:
                row = connection.execute(
                    "SELECT * FROM enterprise_rag_profiles WHERE catalog_id = ? AND profile_name = ?",
                    (self.catalog_id, profile_name),
                ).fetchone()
            else:
                row = connection.execute(
                    """SELECT * FROM enterprise_rag_profiles WHERE catalog_id = ?
                       ORDER BY is_active DESC, updated_at DESC LIMIT 1""", (self.catalog_id,),
                ).fetchone()
        if not row:
            return "default", self.config
        return row["profile_name"], RAGConfig.from_mapping(_parse_json(row["config_json"], {}))

    def import_project_standard(
        self, project_root: str | Path, *, version_label: str = "2018", actor: str = "admin", **kwargs: Any
    ) -> dict[str, Any]:
        source = Path(project_root) / "SY_T5497-2018备品备件分类树(1).xlsx"
        return self.import_xlsx(source, version_label=version_label, actor=actor, expected_counts={
            "entries": 72, "aliases": 12, "references": 10,
        }, **kwargs)

    def import_xlsx(
        self,
        source_path: str | Path,
        *,
        version_label: str,
        actor: str = "admin",
        notes: str = "",
        allowed_plants: Sequence[str] | None = None,
        allowed_classifications: Sequence[str] | None = None,
        security_classification: str = "INTERNAL",
        expected_counts: Mapping[str, int] | None = None,
    ) -> dict[str, Any]:
        """Import an XLSX as a DRAFT version; repeated source hashes are idempotent."""

        path = Path(source_path).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(path)
        if path.suffix.lower() != ".xlsx":
            raise ValueError("only .xlsx standard sources are supported")
        size = path.stat().st_size
        if size <= 0 or size > self.max_source_bytes:
            raise ValueError(f"source size must be between 1 and {self.max_source_bytes} bytes")
        source_sha256 = _sha256_bytes(path.read_bytes())
        label = _clean(version_label)
        if not label:
            raise ValueError("version_label is required")

        with self._connect() as connection:
            existing = connection.execute(
                """SELECT * FROM enterprise_rag_versions
                   WHERE catalog_id = ? AND source_sha256 = ?""", (self.catalog_id, source_sha256),
            ).fetchone()
        if existing:
            return {**self._version_dict(existing), "idempotent": True}

        workbook = self._read_workbook(path)
        parsed = self._parse_standard_workbook(workbook)
        validation = self._validate_parsed(parsed, expected_counts)
        now = _utc_now()
        source_id = f"SRC-{source_sha256[:20].upper()}"
        version_id = f"VER-{uuid.uuid4().hex[:20].upper()}"
        plant_acl = _json_list(allowed_plants)
        classification_acl = _json_list(allowed_classifications, ("INTERNAL", "CONFIDENTIAL", "RESTRICTED"))
        security = _clean(security_classification).upper() or "INTERNAL"

        with self._write() as connection:
            concurrent_existing = connection.execute(
                """SELECT * FROM enterprise_rag_versions
                   WHERE catalog_id = ? AND source_sha256 = ?""", (self.catalog_id, source_sha256),
            ).fetchone()
            if concurrent_existing:
                return {**self._version_dict(concurrent_existing), "idempotent": True}
            conflict = connection.execute(
                """SELECT version_id, source_sha256 FROM enterprise_rag_versions
                   WHERE catalog_id = ? AND version_label = ?""", (self.catalog_id, label),
            ).fetchone()
            if conflict:
                raise ValueError(f"version label already exists with a different source: {label}")
            connection.execute(
                """INSERT OR IGNORE INTO enterprise_rag_sources
                   (source_id, source_sha256, filename, source_path, byte_size, imported_at)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (source_id, source_sha256, path.name, path.name, size, now),
            )
            connection.execute(
                """INSERT INTO enterprise_rag_versions
                   (version_id, catalog_id, version_label, status, source_id, source_sha256,
                    entry_count, alias_count, reference_count, validation_json, allowed_plants_json,
                    allowed_classifications_json, security_classification, notes, created_by, created_at)
                   VALUES (?, ?, ?, 'DRAFT', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (version_id, self.catalog_id, label, source_id, source_sha256,
                 len(parsed["entries"]), len(parsed["aliases"]), len(parsed["references"]),
                 _canonical_json(validation), plant_acl, classification_acl, security,
                 _clean(notes), _clean(actor) or "admin", now),
            )
            self._insert_entries(connection, version_id, parsed["entries"], security, plant_acl, now)
            self._insert_aliases(connection, version_id, parsed["aliases"])
            self._insert_references(connection, version_id, parsed["references"])
            self._append_audit(connection, "STANDARD_VERSION_IMPORTED", version_id, actor, {
                "version_label": label, "source_sha256": source_sha256, "source_filename": path.name,
                "counts": {"entries": len(parsed["entries"]), "aliases": len(parsed["aliases"]),
                           "references": len(parsed["references"])},
                "validation": validation,
            })
            row = connection.execute(
                "SELECT * FROM enterprise_rag_versions WHERE version_id = ?", (version_id,)
            ).fetchone()
        return {**self._version_dict(row), "idempotent": False, "validation": validation}

    def validate_version(self, version_id: str, *, actor: str = "admin") -> dict[str, Any]:
        with self._write() as connection:
            version = self._require_version(connection, version_id)
            if version["status"] not in {STATUS_DRAFT, STATUS_VALIDATED}:
                raise ValueError(f"only DRAFT or VALIDATED versions can be validated, got {version['status']}")
            entries = [dict(row) for row in connection.execute(
                "SELECT * FROM enterprise_rag_entries WHERE version_id = ? ORDER BY source_row", (version_id,)
            )]
            aliases = [dict(row) for row in connection.execute(
                "SELECT * FROM enterprise_rag_aliases WHERE version_id = ? ORDER BY source_row", (version_id,)
            )]
            references = [dict(row) for row in connection.execute(
                "SELECT * FROM enterprise_rag_references WHERE version_id = ? ORDER BY source_row", (version_id,)
            )]
            validation = self._validate_parsed(
                {"entries": entries, "aliases": aliases, "references": references}, None
            )
            connection.execute(
                """UPDATE enterprise_rag_versions SET status = 'VALIDATED', validation_json = ?,
                   validated_at = ? WHERE version_id = ?""",
                (_canonical_json(validation), _utc_now(), version_id),
            )
            self._append_audit(connection, "STANDARD_VERSION_VALIDATED", version_id, actor, validation)
            updated = connection.execute(
                "SELECT * FROM enterprise_rag_versions WHERE version_id = ?", (version_id,)
            ).fetchone()
        return {**self._version_dict(updated), "validation": validation}

    def publish_version(
        self, version_id: str, *, actor: str = "admin", expected_current_version_id: str | None = None
    ) -> dict[str, Any]:
        return self._activate_version(
            version_id, actor=actor, expected_current_version_id=expected_current_version_id,
            event_type="STANDARD_VERSION_PUBLISHED", allow_statuses={STATUS_VALIDATED},
        )

    def rollback_version(
        self, version_id: str, *, actor: str = "admin", expected_current_version_id: str | None = None
    ) -> dict[str, Any]:
        return self._activate_version(
            version_id, actor=actor, expected_current_version_id=expected_current_version_id,
            event_type="STANDARD_VERSION_ROLLED_BACK", allow_statuses={STATUS_RETIRED, STATUS_VALIDATED},
        )

    def _activate_version(
        self, version_id: str, *, actor: str, expected_current_version_id: str | None,
        event_type: str, allow_statuses: set[str],
    ) -> dict[str, Any]:
        with self._write() as connection:
            target = self._require_version(connection, version_id)
            catalog = connection.execute(
                "SELECT * FROM enterprise_rag_catalogs WHERE catalog_id = ?", (self.catalog_id,)
            ).fetchone()
            current_id = catalog["active_version_id"]
            if expected_current_version_id is not None and current_id != expected_current_version_id:
                raise RuntimeError("active version changed; retry with the current version id")
            if current_id == version_id and target["status"] == STATUS_PUBLISHED:
                return {**self._version_dict(target), "idempotent": True, "previous_version_id": current_id}
            if target["status"] not in allow_statuses:
                raise ValueError(f"version status {target['status']} cannot be activated")
            now = _utc_now()
            if current_id:
                connection.execute(
                    """UPDATE enterprise_rag_versions SET status = 'RETIRED', retired_at = ?
                       WHERE version_id = ? AND status = 'PUBLISHED'""", (now, current_id),
                )
            connection.execute(
                """UPDATE enterprise_rag_versions SET status = 'PUBLISHED', published_at = ?, retired_at = NULL
                   WHERE version_id = ?""", (now, version_id),
            )
            connection.execute(
                "UPDATE enterprise_rag_catalogs SET active_version_id = ?, updated_at = ? WHERE catalog_id = ?",
                (version_id, now, self.catalog_id),
            )
            self._append_audit(connection, event_type, version_id, actor, {
                "previous_version_id": current_id, "active_version_id": version_id,
            })
            updated = connection.execute(
                "SELECT * FROM enterprise_rag_versions WHERE version_id = ?", (version_id,)
            ).fetchone()
        return {**self._version_dict(updated), "idempotent": False, "previous_version_id": current_id}

    def list_versions(self) -> list[dict[str, Any]]:
        with self._connect() as connection:
            rows = connection.execute(
                """SELECT * FROM enterprise_rag_versions WHERE catalog_id = ?
                   ORDER BY created_at DESC, version_id DESC""", (self.catalog_id,),
            ).fetchall()
        return [self._version_dict(row) for row in rows]

    def reindex_version(
        self, version_id: str | None = None, *, profile_name: str | None = None, actor: str = "admin"
    ) -> dict[str, Any]:
        """Persist local vectors using a profile's configured dimension."""

        resolved_profile_name, config = self._profile(profile_name)
        with self._write() as connection:
            version = self._resolve_version(connection, version_id, allow_unpublished=True)
            if not version:
                raise LookupError("no standard knowledge version is available")
            chunks = connection.execute(
                """SELECT chunk_id, content, content_hash FROM enterprise_rag_chunks
                   WHERE version_id = ? AND entity_type = 'ENTRY'""", (version["version_id"],),
            ).fetchall()
            now = _utc_now()
            for chunk in chunks:
                vector = self._embedding(chunk["content"], config.dimension)
                connection.execute(
                    """INSERT INTO enterprise_rag_embeddings
                       (chunk_id, provider, model, dimension, vector, content_hash, created_at)
                       VALUES (?, 'local', 'feature-hash-v1', ?, ?, ?, ?)
                       ON CONFLICT(chunk_id, provider, model) DO UPDATE SET
                        dimension=excluded.dimension, vector=excluded.vector,
                        content_hash=excluded.content_hash, created_at=excluded.created_at""",
                    (chunk["chunk_id"], config.dimension, self._pack_vector(vector), chunk["content_hash"], now),
                )
            result = {
                "version_id": version["version_id"], "standard_version": version["version_label"],
                "profile": resolved_profile_name, "indexed": len(chunks), "provider": "local",
                "model": "feature-hash-v1", "dimension": config.dimension,
            }
            self._append_audit(connection, "STANDARD_VECTOR_INDEX_REBUILT", version["version_id"], actor, result)
        return result

    def stats(self, *, version_id: str | None = None) -> dict[str, Any]:
        with self._connect() as connection:
            version = self._resolve_version(connection, version_id, allow_unpublished=bool(version_id))
            version_count = connection.execute(
                "SELECT COUNT(*) FROM enterprise_rag_versions WHERE catalog_id = ?", (self.catalog_id,)
            ).fetchone()[0]
            if not version:
                return {
                    "namespace": "standard_kb", "count": 0, "provider": "local",
                    "model": "feature-hash-v1", "dimension": self.config.dimension,
                    "standard_version": None, "active_version_id": None, "version_count": version_count,
                }
            counts = connection.execute(
                """SELECT COUNT(DISTINCT major_code), COUNT(DISTINCT middle_code),
                          COUNT(DISTINCT minor_code), COUNT(*)
                   FROM enterprise_rag_entries WHERE version_id = ?""", (version["version_id"],),
            ).fetchone()
        return {
            "namespace": "standard_kb", "count": int(counts[3]), "provider": "local",
            "model": "feature-hash-v1", "dimension": self.config.dimension,
            "standard_version": version["version_label"], "active_version_id": version["version_id"],
            "status": version["status"], "version_count": int(version_count),
            "hierarchy": {"major": int(counts[0]), "middle": int(counts[1]), "minor": int(counts[2]),
                          "items": int(counts[3])},
            "aliases": int(version["alias_count"]), "references": int(version["reference_count"]),
            "source_sha256": version["source_sha256"],
        }

    def search(
        self,
        query: str,
        *,
        top_k: int | None = None,
        version_id: str | None = None,
        plant_code: str = "GROUP",
        clearance: str = "INTERNAL",
        profile_name: str | None = None,
        trace_id: str | None = None,
        actor: str | None = None,
        allow_unpublished: bool = False,
    ) -> dict[str, Any]:
        """Hybrid rule + lexical + local-vector search with row-level citations."""

        cleaned_query = _clean(query)
        if not cleaned_query:
            raise ValueError("query is required")
        resolved_profile_name, config = self._profile(profile_name)
        limit = max(1, min(50, int(top_k or config.default_top_k)))
        with self._connect() as connection:
            version = self._resolve_version(connection, version_id, allow_unpublished=allow_unpublished)
            if not version:
                raise LookupError("no published standard knowledge version is available")
            self._assert_access(version, plant_code, clearance)
            entries = [dict(row) for row in connection.execute(
                """SELECT e.*, s.filename, s.source_sha256
                   FROM enterprise_rag_entries e
                   JOIN enterprise_rag_versions v ON v.version_id = e.version_id
                   JOIN enterprise_rag_sources s ON s.source_id = v.source_id
                   WHERE e.version_id = ?""", (version["version_id"],),
            )]
            aliases = [dict(row) for row in connection.execute(
                """SELECT a.*, s.filename, s.source_sha256, v.version_label
                   FROM enterprise_rag_aliases a
                   JOIN enterprise_rag_versions v ON v.version_id = a.version_id
                   JOIN enterprise_rag_sources s ON s.source_id = v.source_id
                   WHERE a.version_id = ?""", (version["version_id"],),
            )]
            embedding_rows = connection.execute(
                """SELECT c.entity_id, e.vector, e.dimension FROM enterprise_rag_chunks c
                   JOIN enterprise_rag_embeddings e ON e.chunk_id = c.chunk_id
                   WHERE c.version_id = ? AND c.entity_type = 'ENTRY'
                     AND e.provider = 'local' AND e.model = 'feature-hash-v1'""", (version["version_id"],),
            ).fetchall()
            reference_rows = [dict(row) for row in connection.execute(
                """SELECT r.*, s.filename, s.source_sha256
                   FROM enterprise_rag_references r
                   JOIN enterprise_rag_versions v ON v.version_id = r.version_id
                   JOIN enterprise_rag_sources s ON s.source_id = v.source_id
                   WHERE r.version_id = ?""", (version["version_id"],),
            )]

        vectors = {row["entity_id"]: self._unpack_vector(row["vector"], row["dimension"]) for row in embedding_rows}
        query_vector = self._embedding(cleaned_query, config.dimension)
        rule_scores, rule_citations, external_aliases = self._rule_candidates(cleaned_query, aliases, entries)
        rule_weight, lexical_weight, vector_weight = config.normalized_weights()
        candidates: list[dict[str, Any]] = []
        for entry in entries:
            lexical = self._lexical_similarity(cleaned_query, entry["content"])
            entry_vector = vectors.get(entry["entry_id"], ())
            if len(entry_vector) != config.dimension:
                entry_vector = self._embedding(entry["content"], config.dimension)
            vector = max(0.0, self._cosine(query_vector, entry_vector))
            rule = rule_scores.get(entry["full_code"], 0.0)
            exact_code = entry["full_code"] in re.sub(r"\D", "", cleaned_query)
            if exact_code:
                rule = 1.0
                lexical = max(lexical, 1.0)
            score = min(1.0, rule * rule_weight + lexical * lexical_weight + vector * vector_weight)
            citations = [self._entry_citation(entry, version)] + rule_citations.get(entry["full_code"], [])
            candidates.append(self._entry_result(
                entry, version, score, rule, lexical, vector, citations, boundary=False,
            ))
        for alias in external_aliases:
            rule = rule_scores.get(alias["target_code"], 0.0)
            lexical = self._lexical_similarity(cleaned_query, f"{alias['keywords']} {alias['level_note']}")
            vector = max(0.0, self._cosine(query_vector, self._embedding(
                f"{alias['keywords']} {alias['major_name']} {alias['level_note']}", config.dimension
            )))
            score = min(1.0, rule * rule_weight + lexical * lexical_weight + vector * vector_weight)
            candidates.append(self._external_result(alias, version, score, rule, lexical, vector))
        candidates.sort(key=lambda item: (item["score"], item["score_breakdown"]["rule"]), reverse=True)
        results = candidates[:limit]
        top_score = results[0]["score"] if results else 0.0
        second_score = results[1]["score"] if len(results) > 1 else 0.0
        margin = round(top_score - second_score, 6)
        decision = "AUTO_RECOMMEND" if (
            top_score >= config.auto_accept_threshold and margin >= config.minimum_margin
        ) else ("REVIEW" if results else "UNRESOLVED")
        reference_evidence = self._rank_reference_evidence(cleaned_query, reference_rows, version, 3)
        payload = {
            "query": cleaned_query, "namespace": "standard_kb", "provider": "local",
            "model": "feature-hash-v1", "dimension": config.dimension,
            "standard_version": version["version_label"], "version_id": version["version_id"],
            "status": version["status"], "results": results, "top_k": limit,
            "total": len(candidates), "citation_required": True,
            "retrieval": {
                "method": "hybrid-rule-lexical-local-vector", "profile": resolved_profile_name,
                "weights": {"rule": rule_weight, "lexical": lexical_weight, "vector": vector_weight},
                "decision": decision, "top_score": round(top_score, 6), "margin": margin,
                "threshold": config.auto_accept_threshold, "minimum_margin": config.minimum_margin,
            },
            "reference_evidence": reference_evidence,
            "acl": {"plant_code": _clean(plant_code).upper() or "GROUP",
                    "clearance": _clean(clearance).upper() or "INTERNAL", "enforced": True},
            "trace_id": _clean(trace_id) or None,
        }
        if config.audit_queries:
            self._record_query(payload, actor=actor)
        return payload

    def classify(
        self,
        material_name: str,
        description: str = "",
        **search_options: Any,
    ) -> dict[str, Any]:
        source_text = _clean(f"{material_name} {description}")
        if not source_text:
            raise ValueError("material_name or description is required")
        search_payload = self.search(source_text, **search_options)
        references = search_payload["results"]
        if not references:
            raise LookupError("standard knowledge search returned no candidates")
        winner = references[0]
        candidates = [{
            "category": item["category"], "score": item["score"],
            "code_prefix": item["code_prefix"], "rag_score": item["score"],
            "full_code": item["full_code"], "reason": "标准知识库混合检索候选",
        } for item in references[:3]]
        decision = search_payload["retrieval"]["decision"]
        return {
            "standard": self.standard_no.replace(" ", ""),
            "recommended_category": winner["category"],
            "recommended_code": winner["full_code"],
            "classification_path": winner["classification_path"],
            "confidence": winner["score"], "code_prefix": winner["code_prefix"],
            "attributes": {}, "standard_name_preview": _clean(material_name) or winner["title"],
            "candidates": candidates, "standard_references": references[:3],
            "evidence_ids": [citation["citation_id"] for citation in winner["citations"]],
            "needs_review": decision != "AUTO_RECOMMEND", "decision_reason": decision,
            "semantic": {
                "method": "本地特征哈希向量 + 规则 + 词法混合检索", "provider": "local",
                "model": "feature-hash-v1", "dimension": search_payload["dimension"],
                "embedding_active": True, "warning": None,
            },
            "rag": {
                "namespace": "standard_kb", "retrieval_count": len(references),
                "method": search_payload["retrieval"]["method"],
                "version_id": search_payload["version_id"],
                "standard_version": search_payload["standard_version"],
                "profile": search_payload["retrieval"]["profile"],
            },
            "plant_code": search_payload["acl"]["plant_code"],
            "trace_id": search_payload["trace_id"],
        }

    def verify_audit_chain(self) -> dict[str, Any]:
        with self._connect() as connection:
            events = connection.execute(
                """SELECT * FROM enterprise_rag_audit_events WHERE catalog_id = ?
                   ORDER BY event_id""", (self.catalog_id,),
            ).fetchall()
        previous_hash = "0" * 64
        errors = []
        for event in events:
            header = {
                "catalog_id": event["catalog_id"], "version_id": event["version_id"],
                "event_type": event["event_type"], "actor": event["actor"],
                "payload_json": event["payload_json"], "previous_hash": event["previous_hash"],
                "created_at": event["created_at"],
            }
            calculated = _sha256_text(_canonical_json(header))
            if event["previous_hash"] != previous_hash:
                errors.append({"event_id": event["event_id"], "error": "previous hash mismatch"})
            if event["event_hash"] != calculated:
                errors.append({"event_id": event["event_id"], "error": "event hash mismatch"})
            previous_hash = event["event_hash"]
        return {
            "catalog_id": self.catalog_id, "valid": not errors, "event_count": len(events),
            "latest_hash": previous_hash if events else None, "errors": errors,
        }

    def _record_query(self, payload: Mapping[str, Any], *, actor: str | None) -> None:
        query_id = f"QRY-{uuid.uuid4().hex[:20].upper()}"
        compact_results = [{
            "full_code": item["full_code"], "score": item["score"],
            "evidence": [citation["citation_id"] for citation in item["citations"]],
        } for item in payload["results"]]
        with self._write() as connection:
            connection.execute(
                """INSERT INTO enterprise_rag_query_logs
                   (query_id, catalog_id, version_id, query_text, query_sha256, plant_code,
                    clearance, profile_name, results_json, trace_id, actor, created_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (query_id, self.catalog_id, payload["version_id"], payload["query"],
                 _sha256_text(payload["query"]), payload["acl"]["plant_code"],
                 payload["acl"]["clearance"], payload["retrieval"]["profile"],
                 _canonical_json(compact_results), payload.get("trace_id"), _clean(actor) or "anonymous", _utc_now()),
            )
            self._append_audit(connection, "RAG_SEARCH", payload["version_id"], actor or "anonymous", {
                "query_id": query_id, "query_sha256": _sha256_text(payload["query"]),
                "plant_code": payload["acl"]["plant_code"], "clearance": payload["acl"]["clearance"],
                "results": compact_results, "trace_id": payload.get("trace_id"),
            })

    def _append_audit(
        self, connection: sqlite3.Connection, event_type: str, version_id: str | None,
        actor: str, payload: Mapping[str, Any],
    ) -> str:
        previous = connection.execute(
            """SELECT event_hash FROM enterprise_rag_audit_events
               WHERE catalog_id = ? ORDER BY event_id DESC LIMIT 1""", (self.catalog_id,),
        ).fetchone()
        previous_hash = previous["event_hash"] if previous else "0" * 64
        payload_json = _canonical_json(payload)
        created_at = _utc_now()
        header = {
            "catalog_id": self.catalog_id, "version_id": version_id,
            "event_type": event_type, "actor": _clean(actor) or "system",
            "payload_json": payload_json, "previous_hash": previous_hash, "created_at": created_at,
        }
        event_hash = _sha256_text(_canonical_json(header))
        connection.execute(
            """INSERT INTO enterprise_rag_audit_events
               (catalog_id, version_id, event_type, actor, payload_json, previous_hash, event_hash, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (self.catalog_id, version_id, event_type, header["actor"], payload_json,
             previous_hash, event_hash, created_at),
        )
        return event_hash

    def _insert_entries(
        self, connection: sqlite3.Connection, version_id: str, entries: Sequence[Mapping[str, Any]],
        security: str, plant_acl: str, created_at: str,
    ) -> None:
        for item in entries:
            entry_id = f"ENT-{version_id[4:12]}-{item['full_code']}"
            content = self._entry_content(item)
            connection.execute(
                """INSERT INTO enterprise_rag_entries
                   (entry_id, version_id, major_code, major_name, middle_code, middle_name,
                    minor_code, minor_name, item_code, item_name, business_type, full_code,
                    example, content, content_hash, source_sheet, source_row, source_range,
                    row_sha256, security_classification, plant_acl_json)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (entry_id, version_id, item["major_code"], item["major_name"], item["middle_code"],
                 item["middle_name"], item["minor_code"], item["minor_name"], item["item_code"],
                 item["item_name"], item["business_type"], item["full_code"], item["example"],
                 content, _sha256_text(content), item["source_sheet"], item["source_row"],
                 item["source_range"], item["row_sha256"], security, plant_acl),
            )
            metadata = {
                "full_code": item["full_code"], "category": item["minor_name"],
                "title": item["item_name"], "classification_path": self._classification_path(item),
            }
            chunk_id = f"CHK-{entry_id}"
            connection.execute(
                """INSERT INTO enterprise_rag_chunks
                   (chunk_id, version_id, entity_type, entity_id, content, content_hash,
                    source_sheet, source_row, source_range, row_sha256, metadata_json,
                    security_classification, plant_acl_json)
                   VALUES (?, ?, 'ENTRY', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (chunk_id, version_id, entry_id, content, _sha256_text(content), item["source_sheet"],
                 item["source_row"], item["source_range"], item["row_sha256"],
                 _canonical_json(metadata), security, plant_acl),
            )
            vector = self._embedding(content, self.config.dimension)
            connection.execute(
                """INSERT INTO enterprise_rag_embeddings
                   (chunk_id, provider, model, dimension, vector, content_hash, created_at)
                   VALUES (?, 'local', 'feature-hash-v1', ?, ?, ?, ?)""",
                (chunk_id, self.config.dimension, self._pack_vector(vector), _sha256_text(content), created_at),
            )

    @staticmethod
    def _insert_aliases(
        connection: sqlite3.Connection, version_id: str, aliases: Sequence[Mapping[str, Any]]
    ) -> None:
        for item in aliases:
            alias_id = f"ALS-{version_id[4:12]}-{item['source_row']:04d}"
            connection.execute(
                """INSERT INTO enterprise_rag_aliases
                   (alias_id, version_id, keywords, major_code, major_name, target_code, level_note,
                    target_scope, source_sheet, source_row, source_range, row_sha256)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (alias_id, version_id, item["keywords"], item["major_code"], item["major_name"],
                 item["target_code"], item["level_note"], item["target_scope"], item["source_sheet"],
                 item["source_row"], item["source_range"], item["row_sha256"]),
            )

    @staticmethod
    def _insert_references(
        connection: sqlite3.Connection, version_id: str, references: Sequence[Mapping[str, Any]]
    ) -> None:
        version = connection.execute(
            "SELECT security_classification, allowed_plants_json FROM enterprise_rag_versions WHERE version_id = ?",
            (version_id,),
        ).fetchone()
        for item in references:
            reference_id = f"REF-{version_id[4:12]}-{item['source_row']:04d}"
            connection.execute(
                """INSERT INTO enterprise_rag_references
                   (reference_id, version_id, scenario, standard_no, standard_name, purpose,
                    source_sheet, source_row, source_range, row_sha256)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (reference_id, version_id, item["scenario"], item["standard_no"], item["standard_name"],
                 item["purpose"], item["source_sheet"], item["source_row"], item["source_range"],
                 item["row_sha256"]),
            )
            content = " ".join((item["scenario"], item["standard_no"], item["standard_name"], item["purpose"]))
            connection.execute(
                """INSERT INTO enterprise_rag_chunks
                   (chunk_id, version_id, entity_type, entity_id, content, content_hash,
                    source_sheet, source_row, source_range, row_sha256, metadata_json,
                    security_classification, plant_acl_json)
                   VALUES (?, ?, 'REFERENCE', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (f"CHK-{reference_id}", version_id, reference_id, content, _sha256_text(content),
                 item["source_sheet"], item["source_row"], item["source_range"], item["row_sha256"],
                 _canonical_json({"standard_no": item["standard_no"], "scenario": item["scenario"]}),
                 version["security_classification"], version["allowed_plants_json"]),
            )

    def _parse_standard_workbook(self, workbook: Mapping[str, list[list[str]]]) -> dict[str, list[dict[str, Any]]]:
        for name, headers in (
            (CLASSIFICATION_SHEET, CLASSIFICATION_HEADERS),
            (MAPPING_SHEET, MAPPING_HEADERS),
            (REFERENCE_SHEET, REFERENCE_HEADERS),
        ):
            if name not in workbook:
                raise ValueError(f"required worksheet is missing: {name}")
            actual = tuple(_clean(value) for value in (workbook[name][0] if workbook[name] else []))
            if actual[:len(headers)] != headers:
                raise ValueError(f"worksheet {name} has unexpected columns: {actual}")

        entries = []
        for row_number, raw in enumerate(workbook[CLASSIFICATION_SHEET][1:], 2):
            values = [_clean(value) for value in (raw + [""] * len(CLASSIFICATION_HEADERS))[:len(CLASSIFICATION_HEADERS)]]
            if not any(values):
                continue
            item = dict(zip((
                "major_code", "major_name", "middle_code", "middle_name", "minor_code", "minor_name",
                "item_code", "item_name", "business_type", "full_code", "example",
            ), values))
            item.update(self._source_fields(CLASSIFICATION_SHEET, row_number, len(CLASSIFICATION_HEADERS), values))
            entries.append(item)

        codes = {item["full_code"] for item in entries}
        aliases = []
        for row_number, raw in enumerate(workbook[MAPPING_SHEET][1:], 2):
            values = [_clean(value) for value in (raw + [""] * len(MAPPING_HEADERS))[:len(MAPPING_HEADERS)]]
            if not any(values):
                continue
            item = dict(zip(("keywords", "major_code", "major_name", "target_code", "level_note"), values))
            item["target_scope"] = "IN_TREE" if item["target_code"] in codes else "EXTERNAL_BOUNDARY"
            item.update(self._source_fields(MAPPING_SHEET, row_number, len(MAPPING_HEADERS), values))
            aliases.append(item)

        references = []
        for row_number, raw in enumerate(workbook[REFERENCE_SHEET][1:], 2):
            values = [_clean(value) for value in (raw + [""] * len(REFERENCE_HEADERS))[:len(REFERENCE_HEADERS)]]
            if not any(values):
                continue
            item = dict(zip(("scenario", "standard_no", "standard_name", "purpose"), values))
            item.update(self._source_fields(REFERENCE_SHEET, row_number, len(REFERENCE_HEADERS), values))
            references.append(item)
        return {"entries": entries, "aliases": aliases, "references": references}

    @staticmethod
    def _source_fields(sheet: str, row_number: int, width: int, values: Sequence[str]) -> dict[str, Any]:
        return {
            "source_sheet": sheet, "source_row": row_number,
            "source_range": f"A{row_number}:{_column_name(width)}{row_number}",
            "row_sha256": _sha256_text(_canonical_json(list(values))),
        }

    @staticmethod
    def _validate_parsed(
        parsed: Mapping[str, Sequence[Mapping[str, Any]]], expected_counts: Mapping[str, int] | None
    ) -> dict[str, Any]:
        entries = list(parsed.get("entries", ()))
        aliases = list(parsed.get("aliases", ()))
        references = list(parsed.get("references", ()))
        errors = []
        seen_codes: set[str] = set()
        for item in entries:
            required = (
                "major_code", "major_name", "middle_code", "middle_name", "minor_code", "minor_name",
                "item_code", "item_name", "business_type", "full_code", "example",
            )
            missing = [key for key in required if not _clean(item.get(key))]
            if missing:
                errors.append({"row": item.get("source_row"), "error": "missing fields", "fields": missing})
                continue
            codes = [item["major_code"], item["middle_code"], item["minor_code"], item["item_code"], item["full_code"]]
            if not all(_CODE_RE.fullmatch(_clean(code)) for code in codes):
                errors.append({"row": item.get("source_row"), "error": "codes must be numeric"})
            if list(map(len, codes)) != [2, 4, 6, 8, 8]:
                errors.append({"row": item.get("source_row"), "error": "invalid code lengths"})
            if not (item["middle_code"].startswith(item["major_code"])
                    and item["minor_code"].startswith(item["middle_code"])
                    and item["item_code"].startswith(item["minor_code"])
                    and item["item_code"] == item["full_code"]):
                errors.append({"row": item.get("source_row"), "error": "hierarchy prefix mismatch"})
            if item["full_code"] in seen_codes:
                errors.append({"row": item.get("source_row"), "error": "duplicate full code"})
            seen_codes.add(item["full_code"])
        for alias in aliases:
            if not all(_clean(alias.get(key)) for key in ("keywords", "major_code", "major_name", "target_code", "level_note")):
                errors.append({"row": alias.get("source_row"), "error": "incomplete alias mapping"})
            if not re.fullmatch(r"\d{8}", _clean(alias.get("target_code"))):
                errors.append({"row": alias.get("source_row"), "error": "alias target must be an 8-digit code"})
        for reference in references:
            if not all(_clean(reference.get(key)) for key in ("scenario", "standard_no", "standard_name", "purpose")):
                errors.append({"row": reference.get("source_row"), "error": "incomplete standard reference"})
        counts = {"entries": len(entries), "aliases": len(aliases), "references": len(references)}
        for key, expected in (expected_counts or {}).items():
            if counts.get(key) != int(expected):
                errors.append({"error": f"expected {expected} {key}, got {counts.get(key)}"})
        if not entries:
            errors.append({"error": "classification tree is empty"})
        if errors:
            raise ValueError(f"standard knowledge validation failed: {_canonical_json(errors[:20])}")
        return {
            "valid": True, "counts": counts,
            "hierarchy": {
                "major": len({item["major_code"] for item in entries}),
                "middle": len({item["middle_code"] for item in entries}),
                "minor": len({item["minor_code"] for item in entries}),
                "items": len(seen_codes),
            },
            "external_aliases": sum(alias.get("target_scope") == "EXTERNAL_BOUNDARY" for alias in aliases),
            "validated_at": _utc_now(),
        }

    def _read_workbook(self, path: Path) -> dict[str, list[list[str]]]:
        self._validate_zip(path)
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            return self._read_workbook_ooxml(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            result = {}
            for sheet in workbook.worksheets:
                result[sheet.title] = [[_clean(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            return result
        finally:
            workbook.close()

    def _validate_zip(self, path: Path) -> None:
        try:
            with zipfile.ZipFile(path) as archive:
                total = 0
                for item in archive.infolist():
                    normalized = posixpath.normpath(item.filename.replace("\\", "/"))
                    if normalized.startswith("../") or normalized.startswith("/"):
                        raise ValueError("unsafe path in XLSX archive")
                    total += int(item.file_size)
                    if total > self.max_uncompressed_bytes:
                        raise ValueError("XLSX uncompressed content exceeds configured limit")
        except zipfile.BadZipFile as exc:
            raise ValueError("invalid XLSX archive") from exc

    @staticmethod
    def _read_workbook_ooxml(path: Path) -> dict[str, list[list[str]]]:
        with zipfile.ZipFile(path) as archive:
            shared: list[str] = []
            if "xl/sharedStrings.xml" in archive.namelist():
                root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                for item in root.findall(f"{{{_MAIN_NS}}}si"):
                    shared.append("".join(node.text or "" for node in item.iter(f"{{{_MAIN_NS}}}t")))
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            targets = {
                rel.attrib["Id"]: rel.attrib["Target"]
                for rel in relationships.findall(f"{{{_PKG_REL_NS}}}Relationship")
            }
            result: dict[str, list[list[str]]] = {}
            sheets = workbook.find(f"{{{_MAIN_NS}}}sheets")
            if sheets is None:
                raise ValueError("XLSX workbook has no sheets")
            for sheet in sheets:
                name = sheet.attrib["name"]
                relation_id = sheet.attrib[f"{{{_REL_NS}}}id"]
                target = targets.get(relation_id, "")
                entry = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
                if not entry.startswith("xl/") or entry not in archive.namelist():
                    raise ValueError(f"invalid worksheet target for {name}")
                root = ET.fromstring(archive.read(entry))
                rows: list[list[str]] = []
                sheet_data = root.find(f"{{{_MAIN_NS}}}sheetData")
                if sheet_data is None:
                    result[name] = rows
                    continue
                for row_node in sheet_data.findall(f"{{{_MAIN_NS}}}row"):
                    values: dict[int, str] = {}
                    for cell in row_node.findall(f"{{{_MAIN_NS}}}c"):
                        reference = cell.attrib.get("r", "")
                        column = _column_number(reference)
                        cell_type = cell.attrib.get("t")
                        value_node = cell.find(f"{{{_MAIN_NS}}}v")
                        if cell_type == "inlineStr":
                            value = "".join(node.text or "" for node in cell.iter(f"{{{_MAIN_NS}}}t"))
                        elif value_node is None:
                            value = ""
                        elif cell_type == "s":
                            index = int(value_node.text or "0")
                            value = shared[index] if 0 <= index < len(shared) else ""
                        elif cell_type == "b":
                            value = "TRUE" if value_node.text == "1" else "FALSE"
                        else:
                            value = value_node.text or ""
                        values[column] = _clean(value)
                    width = max(values, default=0)
                    rows.append([values.get(index, "") for index in range(1, width + 1)])
                result[name] = rows
            return result

    def _embedding(self, text: str, dimension: int) -> tuple[float, ...]:
        if self.embedding_fn is not None:
            values = tuple(float(value) for value in self.embedding_fn(text, dimension))
            if len(values) != dimension or any(not math.isfinite(value) for value in values):
                raise ValueError("embedding_fn returned an invalid vector")
            return self._normalize_vector(values)
        normalized = re.sub(r"\s+", "", _clean(text).lower())
        tokens = list(normalized)
        tokens.extend(normalized[index:index + 2] for index in range(max(0, len(normalized) - 1)))
        tokens.extend(_ALNUM_RE.findall(normalized))
        vector = [0.0] * dimension
        for token in tokens or ["<empty>"]:
            digest = hashlib.sha256(token.encode("utf-8")).digest()
            index = int.from_bytes(digest[:4], "big") % dimension
            vector[index] += 1.0 if digest[4] & 1 else -1.0
        return self._normalize_vector(vector)

    @staticmethod
    def _normalize_vector(vector: Iterable[float]) -> tuple[float, ...]:
        values = tuple(float(value) for value in vector)
        norm = math.sqrt(sum(value * value for value in values))
        return tuple(value / norm for value in values) if norm else values

    @staticmethod
    def _pack_vector(vector: Sequence[float]) -> bytes:
        return struct.pack(f"<{len(vector)}f", *vector)

    @staticmethod
    def _unpack_vector(value: bytes, dimension: int) -> tuple[float, ...]:
        if not value or len(value) != int(dimension) * 4:
            return ()
        return struct.unpack(f"<{int(dimension)}f", value)

    @staticmethod
    def _cosine(left: Sequence[float], right: Sequence[float]) -> float:
        if not left or len(left) != len(right):
            return 0.0
        denominator = math.sqrt(sum(value * value for value in left)) * math.sqrt(sum(value * value for value in right))
        return sum(a * b for a, b in zip(left, right)) / denominator if denominator else 0.0

    @staticmethod
    def _terms(text: str) -> set[str]:
        normalized = re.sub(r"\s+", "", _clean(text).lower())
        terms = set(normalized)
        terms.update(normalized[index:index + 2] for index in range(max(0, len(normalized) - 1)))
        terms.update(_ALNUM_RE.findall(normalized))
        return {term for term in terms if term}

    @classmethod
    def _lexical_similarity(cls, query: str, content: str) -> float:
        query_terms, content_terms = cls._terms(query), cls._terms(content)
        if not query_terms or not content_terms:
            return 0.0
        overlap = len(query_terms & content_terms) / len(query_terms)
        query_normalized = re.sub(r"\s+", "", query.lower())
        content_normalized = re.sub(r"\s+", "", content.lower())
        phrase_bonus = 0.25 if query_normalized and query_normalized in content_normalized else 0.0
        return min(1.0, overlap * 0.75 + phrase_bonus)

    def _rule_candidates(
        self, query: str, aliases: Sequence[Mapping[str, Any]], entries: Sequence[Mapping[str, Any]]
    ) -> tuple[dict[str, float], dict[str, list[dict[str, Any]]], list[Mapping[str, Any]]]:
        normalized_query = re.sub(r"\s+", "", query.lower())
        entry_codes = {entry["full_code"] for entry in entries}
        scores: dict[str, float] = {}
        citations: dict[str, list[dict[str, Any]]] = {}
        external = []
        for alias in aliases:
            phrases = [re.sub(r"\s+", "", phrase.lower()) for phrase in _SPLIT_RE.split(alias["keywords"]) if phrase]
            exact = [phrase for phrase in phrases if phrase and phrase in normalized_query]
            token_score = self._lexical_similarity(query, alias["keywords"])
            rule_score = 1.0 if exact else (0.65 if token_score >= 0.6 else token_score * 0.5)
            if rule_score <= 0:
                continue
            target = alias["target_code"]
            scores[target] = max(scores.get(target, 0.0), rule_score)
            citation = self._alias_citation(alias)
            citations.setdefault(target, []).append(citation)
            if target not in entry_codes:
                external.append(alias)
        unique_external = {item["target_code"]: item for item in external}
        return scores, citations, list(unique_external.values())

    def _entry_result(
        self, entry: Mapping[str, Any], version: Mapping[str, Any], score: float,
        rule: float, lexical: float, vector: float, citations: list[dict[str, Any]], boundary: bool,
    ) -> dict[str, Any]:
        return {
            "reference_id": f"SYT-{entry['full_code']}", "score": round(score, 6),
            "content": entry["content"], "category": entry["minor_name"],
            "code_prefix": f"MDM-{entry['major_code']}", "title": entry["item_name"],
            "keywords": entry["example"], "source": self.standard_no,
            "standard_version": version["version_label"], "version_id": version["version_id"],
            "full_code": entry["full_code"],
            "classification_path": self._classification_path(entry),
            "business_type": entry["business_type"], "boundary_rule": boundary,
            "score_breakdown": {"rule": round(rule, 6), "lexical": round(lexical, 6),
                                "vector": round(vector, 6)},
            "citations": self._deduplicate_citations(citations),
        }

    def _external_result(
        self, alias: Mapping[str, Any], version: Mapping[str, Any], score: float,
        rule: float, lexical: float, vector: float,
    ) -> dict[str, Any]:
        return {
            "reference_id": f"BOUNDARY-{alias['target_code']}", "score": round(score, 6),
            "content": f"{alias['keywords']} {alias['level_note']}", "category": alias["major_name"],
            "code_prefix": f"MDM-{alias['major_code']}", "title": alias["keywords"],
            "keywords": alias["keywords"], "source": self.standard_no,
            "standard_version": version["version_label"], "version_id": version["version_id"],
            "full_code": alias["target_code"],
            "classification_path": f"{alias['major_code']} {alias['major_name']} / {alias['level_note']}",
            "business_type": "整机设备边界", "boundary_rule": True,
            "warning": "该规则属于完整设备边界，不在当前备品备件72条分类树内。",
            "score_breakdown": {"rule": round(rule, 6), "lexical": round(lexical, 6),
                                "vector": round(vector, 6)},
            "citations": [self._alias_citation(alias)],
        }

    @staticmethod
    def _classification_path(item: Mapping[str, Any]) -> str:
        return " / ".join((
            f"{item['major_code']} {item['major_name']}", f"{item['middle_code']} {item['middle_name']}",
            f"{item['minor_code']} {item['minor_name']}", f"{item['full_code']} {item['item_name']}",
        ))

    @classmethod
    def _entry_content(cls, item: Mapping[str, Any]) -> str:
        return " ".join((cls._classification_path(item), item["business_type"], item["example"]))

    def _entry_citation(self, entry: Mapping[str, Any], version: Mapping[str, Any]) -> dict[str, Any]:
        return {
            "citation_id": self._citation_id(entry["source_sha256"], entry["source_sheet"], entry["source_row"]),
            "standard_no": self.standard_no, "version_label": version["version_label"],
            "filename": entry["filename"], "source_sha256": entry["source_sha256"],
            "sheet": entry["source_sheet"], "row": int(entry["source_row"]),
            "range": entry["source_range"], "record_sha256": entry["row_sha256"],
            "excerpt": f"{entry['full_code']} {entry['item_name']}；示例：{entry['example']}",
        }

    def _alias_citation(self, alias: Mapping[str, Any]) -> dict[str, Any]:
        if all(alias.get(key) for key in ("filename", "source_sha256", "version_label")):
            source = alias
        else:
            with self._connect() as connection:
                source = connection.execute(
                    """SELECT s.filename, s.source_sha256, v.version_label
                       FROM enterprise_rag_versions v JOIN enterprise_rag_sources s ON s.source_id = v.source_id
                       WHERE v.version_id = ?""", (alias["version_id"],),
                ).fetchone()
        return {
            "citation_id": self._citation_id(source["source_sha256"], alias["source_sheet"], alias["source_row"]),
            "standard_no": self.standard_no, "version_label": source["version_label"],
            "filename": source["filename"], "source_sha256": source["source_sha256"],
            "sheet": alias["source_sheet"], "row": int(alias["source_row"]),
            "range": alias["source_range"], "record_sha256": alias["row_sha256"],
            "excerpt": f"{alias['keywords']} → {alias['target_code']}（{alias['level_note']}）",
        }

    def _rank_reference_evidence(
        self, query: str, rows: Sequence[Mapping[str, Any]], version: Mapping[str, Any], top_k: int
    ) -> list[dict[str, Any]]:
        ranked = []
        for row in rows:
            content = " ".join((row["scenario"], row["standard_no"], row["standard_name"], row["purpose"]))
            score = self._lexical_similarity(query, content)
            ranked.append({
                "reference_id": row["reference_id"], "score": round(score, 6),
                "scenario": row["scenario"], "standard_no": row["standard_no"],
                "standard_name": row["standard_name"], "purpose": row["purpose"],
                "citation": {
                    "citation_id": self._citation_id(row["source_sha256"], row["source_sheet"], row["source_row"]),
                    "version_label": version["version_label"], "filename": row["filename"],
                    "source_sha256": row["source_sha256"], "sheet": row["source_sheet"],
                    "row": int(row["source_row"]), "range": row["source_range"],
                    "record_sha256": row["row_sha256"],
                },
            })
        ranked.sort(key=lambda item: item["score"], reverse=True)
        return ranked[:top_k]

    @staticmethod
    def _citation_id(source_hash: str, sheet: str, row: int) -> str:
        return f"{source_hash[:12].upper()}#{sheet}!{int(row)}"

    @staticmethod
    def _deduplicate_citations(citations: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
        return list({item["citation_id"]: dict(item) for item in citations}.values())

    @staticmethod
    def _assert_access(version: Mapping[str, Any], plant_code: str, clearance: str) -> None:
        plant = _clean(plant_code).upper() or "GROUP"
        level = _clean(clearance).upper() or "INTERNAL"
        allowed_plants = set(_parse_json(version["allowed_plants_json"], ["*"]))
        allowed_levels = set(_parse_json(version["allowed_classifications_json"], ["INTERNAL"]))
        if "*" not in allowed_plants and plant not in allowed_plants:
            raise PermissionError(f"plant {plant} cannot access this knowledge version")
        if "*" not in allowed_levels and level not in allowed_levels:
            raise PermissionError(f"clearance {level} cannot access this knowledge version")

    def _resolve_version(
        self, connection: sqlite3.Connection, version_id: str | None, *, allow_unpublished: bool
    ) -> sqlite3.Row | None:
        if version_id:
            row = connection.execute(
                "SELECT * FROM enterprise_rag_versions WHERE catalog_id = ? AND version_id = ?",
                (self.catalog_id, version_id),
            ).fetchone()
            if row and row["status"] != STATUS_PUBLISHED and not allow_unpublished:
                raise PermissionError("only published knowledge versions are searchable")
            return row
        return connection.execute(
            """SELECT v.* FROM enterprise_rag_catalogs c
               JOIN enterprise_rag_versions v ON v.version_id = c.active_version_id
               WHERE c.catalog_id = ? AND v.status = 'PUBLISHED'""", (self.catalog_id,),
        ).fetchone()

    def _require_version(self, connection: sqlite3.Connection, version_id: str) -> sqlite3.Row:
        row = connection.execute(
            "SELECT * FROM enterprise_rag_versions WHERE catalog_id = ? AND version_id = ?",
            (self.catalog_id, _clean(version_id)),
        ).fetchone()
        if not row:
            raise LookupError(f"knowledge version not found: {version_id}")
        return row

    @staticmethod
    def _version_dict(row: Mapping[str, Any]) -> dict[str, Any]:
        return {
            "version_id": row["version_id"], "catalog_id": row["catalog_id"],
            "version_label": row["version_label"], "status": row["status"],
            "source_id": row["source_id"], "source_sha256": row["source_sha256"],
            "entry_count": int(row["entry_count"]), "alias_count": int(row["alias_count"]),
            "reference_count": int(row["reference_count"]),
            "validation": _parse_json(row["validation_json"], None),
            "allowed_plants": _parse_json(row["allowed_plants_json"], []),
            "allowed_classifications": _parse_json(row["allowed_classifications_json"], []),
            "security_classification": row["security_classification"], "notes": row["notes"],
            "created_by": row["created_by"], "created_at": row["created_at"],
            "validated_at": row["validated_at"], "published_at": row["published_at"],
            "retired_at": row["retired_at"],
        }


__all__ = [
    "EnterpriseRAG", "RAGConfig", "STATUS_DRAFT", "STATUS_VALIDATED",
    "STATUS_PUBLISHED", "STATUS_RETIRED",
]
